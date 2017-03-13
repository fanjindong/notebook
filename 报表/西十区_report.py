from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import colors, Font, Border, Side, PatternFill, Alignment, Protection

import json
import copy
from pymongo import MongoClient
import logging
import time
from datetime import datetime
from pprint import pprint
import redis


class Xishiqu(object):
    abstract = True
    _mdb = None
    _rdb1 = None
    _rdb2 = None

    @property
    def mdb(self):
        if self._mdb is None:
            _mdb = MongoClient(
                'mongodb://mongoc:Boluome123@139.198.1.168:10018/')
            self._mdb = _mdb["boluome"]
        return self._mdb

    @property
    def rdb_1(self):
        if self._rdb1 is None:
            self._rdb1 = redis.StrictRedis(host='192.168.1.11',
                                           port='6379',
                                           db=1, charset="utf-8",
                                           decode_responses=True)
        return self._rdb1

    @property
    def rdb_2(self):
        if self._rdb2 is None:
            self._rdb2 = redis.StrictRedis(host='192.168.1.11',
                                           port='6379',
                                           db=2, charset="utf-8",
                                           decode_responses=True)
        return self._rdb2

x = Xishiqu()


def mongodb_datas_fetch(start, end, datas=None):
    start = start.split('-')
    end = end.split('-')
    starttime = time.mktime(
        (int(start[0]), int(start[1]), int(start[2]), 0, 0, 0, 0, 0, 0)) * 1000
    endtime = time.mktime(
        (int(end[0]), int(end[1]), int(end[2]), 24, 0, 0, 0, 0, 0)) * 1000

    if datas is None:
        datas = []
    # 查询mongo数据库
    items = x.mdb['order_piaowu'].find(
        {'createdAt': {'$gte': starttime, '$lt': endtime}}, {'_id': 0})
    for item in items:
        order_type = '票务'
        order_id = item['id']
        order_name = item['name']
        order_channel = item['channel']
        order_phone = item.get('userPhone')
        price = item['price']
        xdtime = str(datetime.fromtimestamp(
            item['createdAt'] / 1000)).split('.')[0]
        order_status = item['displayStatus']

        data = {
            'order_type': order_type,
            'order_id': order_id,
            'order_channel': order_channel,
            'order_phone': order_phone,
            'order_name': order_name,
            'xdtime': xdtime,
            'order_status': order_status,
            'order_price': price,
        }
        datas.append(data)
    return datas


def map_values(report_data):

    map_order_channel = {
        'xishiqu': '西十区',
    }

    report_data['order_channel'] = map_order_channel[
        report_data['order_channel'].lower()]
    return report_data


def map_keys(report_data):
    map_key = {
        'order_id': '订单号',
        'order_type': '品类',
        'order_channel': '服务商',
        'order_phone': '手机号',
        'order_name': '订单名称',
        'order_price': '订单金额',
        'order_status': '订单状态',
        'xdtime': '下单时间',
    }
    report_data = {map_key[k]: v for k,
                   v in report_data.items() if k in map_key}
    return report_data


def generate_report(report_datas, field_header):
    wb = Workbook()
    ws = wb.active
    ws.append(field_header)

    # 动态创建各分工作表
    items = copy.deepcopy(report_datas)
    for item in items:
        item = map_values(item)
        item = map_keys(item)
        row = ws.max_row + 1
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = item[
                ws.cell(row=1, column=col).value]

    render_report(ws)
    print('ok')
    # 报表完成，保存为excel文件
    wb.save("/home/fanjindong/notebook/报表/西十区_{}_对账单.xlsx".format('2月'))
    return ['file save is ok']


def render_report(ws):
    # 冻结第一行
    ws.freeze_panes = 'A2'
    for i in range(1, ws.max_column + 1):
        # 抬头第一行 设置填充色、字体格式、居中
        ws['{}1'.format(get_column_letter(i))].fill = PatternFill(
            "solid", fgColor="FEC000")
        ws['{}1'.format(get_column_letter(i))].font = Font(
            name="微软雅黑", size=12, bold=True)
        ws['{}1'.format(get_column_letter(i))].alignment = Alignment(
            horizontal='center', vertical='center')
    # 为 所有表格设置边框和字体。
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom)  # 设置单元格边框格式
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).font = Font(name="微软雅黑", size=11)


if __name__ == "__main__":

    datas = mongodb_datas_fetch('2017-02-01', '2017-02-28')
    # print(datas)
    field_header = ['订单号', '订单名称', '品类', '服务商', '手机号', '订单金额', '下单时间', '订单状态']
    generate_report(datas, field_header)
