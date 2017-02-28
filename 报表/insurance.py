from openpyxl import Workbook
import openpyxl
from pymongo import MongoClient
import pymongo
import time
import json


class Insurance():
    abstract = True
    _mdb = None

    @property
    def mdb(self):
        if self._mdb is None:
            _mdb = MongoClient(
                'mongodb://mongoc:Boluome123@139.198.1.168:10017/')
            self._mdb = _mdb['boluome']
        return self._mdb

isr = Insurance()


def insurance_data_find_from_mongdb(starttime, endtime=time.time()):
    """starttime='2017-01-01,endtime='2017-03-01'
       返回查询到的数据集datas
    """
    datas = []  # 存放所有数据
    ts = time.strptime(starttime, "%Y-%m-%d")
    starttime = time.mktime(ts)
    # print(starttime)
    jipiao_datas = isr.mdb['order_jipiao'].find(
        {"createdAt": {"$gt": starttime * 1000},
         "insurance": {"$ne": None}}, {"_id": 0}
    ).sort("createdAt", pymongo.ASCENDING)
    huoche_datas = isr.mdb['order_huoche'].find(
        {"createdAt": {"$gt": starttime * 1000},
         "insurance": {"$ne": None}}, {"_id": 0}
    ).sort("createdAt", pymongo.ASCENDING)
    # for data in jipiao_datas:
    #     datas.append(data)
    # for data in huoche_datas:
    #     datas.append(data)
    return list(jipiao_datas), list(huoche_datas)


def datas_insert_to_excel(jipiao_datas, huoche_datas):
    wb = Workbook()
    ws = wb.create_sheet(title="jipiao_datas", index=0)
    sub = set()
    for data in jipiao_datas:
        for key in data.keys():
            sub.add(key)
    print(sub, len(sub))
    ws.append(list(sub))  # 第一行
    # 插入数据，从第二行开始
    for row, data in enumerate(jipiao_datas, 2):
        for col in range(1, ws.max_column + 1):
            col = openpyxl.utils.get_column_letter(col)
            key = ws["{}1".format(col)].value
            if data.get(str(key), ''):
                ws["{}{}".format(col, row)].value = json.dumps(
                    data.get(str(key), ''), indent=2, ensure_ascii=False)

    ws = wb.create_sheet(title="huoche_datas", index=0)
    sub = set()
    for data in huoche_datas:
        for key in data.keys():
            sub.add(key)
    print(sub, len(sub))
    ws.append(list(sub))  # 第一行
    # 插入数据，从第二行开始
    for row, data in enumerate(huoche_datas, 2):
        for col in range(1, ws.max_column + 1):
            col = openpyxl.utils.get_column_letter(col)
            key = ws["{}1".format(col)].value
            if data.get(str(key), ''):
                ws["{}{}".format(col, row)].value = json.dumps(
                    data.get(str(key), ''), indent=2, ensure_ascii=False)
    wb.save("/home/fanjindong/notebook/报表/insurance.xlsx")


if __name__ == "__main__":
    jipiao_datas, huoche_datas = insurance_data_find_from_mongdb('2016-10-01')

    datas_insert_to_excel(jipiao_datas, huoche_datas)
    print('ok')
