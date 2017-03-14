from openpyxl import Workbook
from pymongo import MongoClient
import pymongo
import time
from pprint import pprint
import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import colors, Font, Border, Side, PatternFill, Alignment, Protection


class Redenvelopes():
    abstract = True
    _dev = None

    @property
    def dev(self):
        if self._dev is None:
            _dev = MongoClient(
                'mongodb://root:Boluome123@139.198.1.168:11017/')
            self._dev = _dev["boluome"]
        return self._dev

ree = Redenvelopes()


def normal_datas_fetch(yesterday):
    """获取主(normal)红包活动数据(默认昨日)"""

    normal_datas = ree.dev['activity'].find(
        {"target": "normal", "date": yesterday}
    )
    return normal_datas


def coupon_datas_fetch(activity_id, start, end):
    """获取activityId对应的副(coupon)红包活动数据(start-endtime)"""
    coupon_datas = []
    for coupon_data in ree.dev['activity'].find({"activityId": activity_id, "target": "coupon", "date": {"$gte": start, "$lte": end}}):

        if coupon_data['giveType'] == 1.0:
            coupon_data['giveType'] = '固定投放'
        else:
            coupon_data['giveType'] = '随机投放'

        coupon_data['unused'] = coupon_data[
            'distributed'] - coupon_data['used']
        coupon_datas.append(coupon_data)

    return coupon_datas


def normal_datas_to_excel(normal_data, coupon_datas, start, end):
    wb = Workbook()
    ws = wb.active
    ws.title = "优惠活动报告 每月汇总明细"
    ws.append(["红包活动报告-【{}】".format(normal_data['activityName'])] * 4)
    ws.append([''])
    ws.merge_cells("A1:D2")
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = PatternFill(
                patternType="solid", fgColor="56237f"
            )
            ws.cell(row=row, column=col).font = Font(
                name="黑体", size=16, color=colors.WHITE)
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center", vertical="center")
    ws.append([''] * 4)
    ws.append(['', '活动总览', '导出数据'])
    ws.append(['起止日期', '', start + ' - ' + end])
    ws.append(['总数', normal_data['total'], normal_data['total']])
    ws.append(['已领取', normal_data['distributed'], ''])
    ws.append(['未领取', normal_data['undistributed'], ''])
    ws.append(['已使用', normal_data['used'], ''])
    ws.append(['未使用', normal_data['distributed'] - normal_data['used'], ''])
    ws.append(['活动总金额', '', ''])
    ws.append(['已使用金额', '', ''])

    for row in range(4, 13):
        ws['B{}'.format(row)].fill = PatternFill("solid", fgColor="ffe699")
        ws['C{}'.format(row)].fill = PatternFill("solid", fgColor="bdd7ee")

    coupon_datas_to_excel(ws, coupon_datas)

    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    for row in range(4, 13):
        for col in range(1, 4):
            ws.cell(row=row, column=col).font = Font(name='DengXian')
            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom)
    for row in range(18, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=18, column=col).fill = PatternFill(
                "solid", fgColor="deebf7")
            ws.cell(row=row, column=col).font = Font(name='DengXian')
            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom)
    wb.save('/home/fanjindong/notebook/报表/{}-{}-{}_{}.xlsx'.format(
        normal_data['appCode'],
        normal_data['activityId'],
        start, end)
    )


def coupon_datas_to_excel(ws, coupon_datas):
    for row in range(13, 17):
        ws.append([''])
    ws.append(['优惠券明细'] * 2)
    ws.merge_cells("A17:B17")
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    for col in range(1, 3):
        ws.cell(row=17, column=col).alignment = Alignment(
            horizontal="center", vertical="center")
        ws.cell(row=17, column=col).fill = PatternFill(
            patternType="solid", fgColor="9dc3e6")
        ws.cell(row=17, column=col).border = Border(
            left=left, right=right, top=top, bottom=bottom)
        ws.cell(row=17, column=col).font = Font(name='DengXian', size=17)
    eng_chi_dict = {'日期': 'date', '策略方式': 'giveType', '红包': 'couponName',
                    '红包总数': 'total', '已领取': 'distributed', '未领取': 'undistributed',
                    '已使用': 'used', '未使用': 'unused', '今日领取': 'distributedToday',
                    '今日使用': 'usedToday', '补贴金额': ''}
    sublist = ['日期', '策略方式', '红包', '红包总数', '已领取',
               '未领取', '已使用', '未使用', '今日领取', '今日使用', '补贴金额']
    ws.append(sublist)
    for index, coupon_data in enumerate(coupon_datas, 19):
        for col in range(1, len(eng_chi_dict) + 1):
            ws.cell(row=index, column=col).value = coupon_data.get(
                eng_chi_dict[ws.cell(row=18, column=col).value]
            )
    ws['C7'].value = "=SUM(I19:I{})".format(ws.max_row)
    ws['C8'].value = "=B8"
    ws['C9'].value = "=SUM(J19:J{})".format(ws.max_row)
    ws['C10'].value = "=C7-C9"
    ws['C12'].value = "=SUM(K19:K{})".format(ws.max_row)


def red_envelopes(start, end):

    normal_datas = normal_datas_fetch(str(end))

    for normal_data in normal_datas:
        activity_id = normal_data['activityId']
        coupon_datas = coupon_datas_fetch(activity_id, str(start), str(end))
        normal_datas_to_excel(normal_data, coupon_datas, str(start), str(end))

if __name__ == "__main__":
    # today = datetime.date.today()
    # yesterday = today - datetime.timedelta(days=1)
    red_envelopes(start='2017-03-01', end='2017-03-13')
