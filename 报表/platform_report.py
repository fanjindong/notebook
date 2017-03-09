from openpyxl import Workbook
from pymongo import MongoClient
import pymongo
import time
from pprint import pprint
import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import colors, Font, Border, Side, PatternFill, Alignment, Protection


class Platfromenvelopes():
    abstract = True
    _dev = None

    @property
    def dev(self):
        if self._dev is None:
            _dev = MongoClient(
                'mongodb://root:Boluome123@139.198.1.168:11017/')
            self._dev = _dev["boluome"]
        return self._dev

ree = Platfromenvelopes()


def activity_ids_fetch():
    """获取activity_ids"""
    activity_ids = set()
    normal_datas = ree.dev['activity'].find(
        {"target": "platform"}
    )
    for normal_data in normal_datas:
        activity_ids.add(normal_data['activityId'])
    return activity_ids


def platform_datas_fetch(activity_id, start, end):
    """获取activityId对应的副(platform)红包活动数据(start-endtime)"""
    platform_datas = []
    normal_datas = {}
    for platform_data in ree.dev['activity'].find({"activityId": activity_id, "target": "platform", "date": {"$gte": start, "$lte": end}}).sort('date'):
        platform_datas.append(platform_data)
        normal_datas[activity_id] = platform_data

    return normal_datas, platform_datas


def normal_datas_to_excel(normal_data, platform_datas, start, end):
    wb = Workbook()
    ws = wb.active
    ws.title = "平台活动报表"
    ws.append(["平台活动报告"] * 4)
    ws.append([''])
    ws.merge_cells("A1:D2")
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = PatternFill(
                patternType="solid", fgColor="56237f"
            )
            ws.cell(row=row, column=col).font = Font(
                name="黑体", size=16, color=colors.WHITE)
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center", vertical="center")
    ws.append([''] * 4)
    ws.append([''] * 4)
    ws.append(['日期', start + ' - ' + end])
    ws.merge_cells("B5:D5")

    ws.append(['活动名', normal_data['activityName']])
    ws.merge_cells("B6:D6")

    for row in range(5, 7):
        for col in range(2, 5):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center", vertical="center")

    ws.append(['已使用', normal_data['used'],
               '剩余', normal_data['remaining']])

    ws.append(['使用', normal_data['used']])
    ws.merge_cells("B8:D8")
    for col in range(2, 5):
        ws.cell(row=ws.max_row, column=col).alignment = Alignment(
            horizontal="center", vertical="center")

    platform_datas_to_excel(ws, platform_datas)

    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    for row in range(5, 9):
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = Font(name='DengXian')
            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom)
    for row in range(15, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=15, column=col).fill = PatternFill(
                "solid", fgColor="deebf7")
            ws.cell(row=row, column=col).font = Font(name='DengXian')
            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom)
    wb.save('/home/fanjindong/notebook/报表/{}-{}-{}_{}.xlsx'.format(
        normal_data['appCode'],
        normal_data['activityId'],
        start, end)
    )


def platform_datas_to_excel(ws, platform_datas):
    for row in range(9, 14):
        ws.append([''])
    ws.append(['优惠券明细'] * 2)
    ws.merge_cells("A14:B14")
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    for col in range(1, 3):
        ws.cell(row=14, column=col).alignment = Alignment(
            horizontal="center", vertical="center")
        ws.cell(row=14, column=col).fill = PatternFill(
            patternType="solid", fgColor="9dc3e6")
        ws.cell(row=14, column=col).border = Border(
            left=left, right=right, top=top, bottom=bottom)
        ws.cell(row=14, column=col).font = Font(name='DengXian', size=14)
    eng_chi_dict = {'日期': 'date', '总数量': 'total', '已使用': 'used',
                    '剩余': 'remaining', '今日使用': 'usedToday'}
    sublist = ['日期', '总数量', '已使用', '剩余', '今日使用']
    ws.append(sublist)
    for index, platform_data in enumerate(platform_datas, 16):
        for col in range(1, len(eng_chi_dict) + 1):
            ws.cell(row=index, column=col).value = platform_data[
                eng_chi_dict[ws.cell(row=15, column=col).value]
            ]
    ws['B8'].value = "=SUM(E16:E{})".format(ws.max_row)


def platform_envelopes(start, end):

    activity_ids = activity_ids_fetch()

    for activity_id in activity_ids:

        normal_data, platform_datas = platform_datas_fetch(
            activity_id, str(start), str(end))
        normal_datas_to_excel(normal_data[activity_id], platform_datas,
                              str(start), str(end))

if __name__ == "__main__":

    platform_envelopes(start='2017-02-01', end='2017-03-08')
