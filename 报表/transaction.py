from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import colors, Font, Border, Side, PatternFill, Alignment, Protection

import json
import copy
from pymongo import MongoClient
import logging
import time
from datetime import datetime
from pprint import pprint
import redis


class Transaction(object):
    abstract = True
    _mdb = None
    _rdb1 = None
    _rdb2 = None

    @property
    def mdb(self):
        if self._mdb is None:
            _mdb = MongoClient(
                'mongodb://mongoc:Boluome123@139.198.1.168:10018/')
            self._mdb = _mdb["boluome"]
        return self._mdb

    @property
    def rdb_1(self):
        if self._rdb1 is None:
            self._rdb1 = redis.StrictRedis(
                host='192.168.1.11',
                port='6379',
                db=1,
                charset="utf-8",
                decode_responses=True)
        return self._rdb1

    @property
    def rdb_2(self):
        if self._rdb2 is None:
            self._rdb2 = redis.StrictRedis(
                host='192.168.1.11',
                port='6379',
                db=2,
                charset="utf-8",
                decode_responses=True)
        return self._rdb2


b = Transaction()


def parameter_verification(appcode, start, end):
    """参数验证"""
    try:
        start = start.split('-')
        end = end.split('-')
        starttime = time.mktime((int(start[0]), int(start[1]), int(start[2]),
                                 0, 0, 0, 0, 0, 0)) * 1000
        endtime = time.mktime(
            (int(end[0]), int(end[1]), int(end[2]), 24, 0, 0, 0, 0, 0)) * 1000
    except ValueError as e:
        print('时间格式错误!')
        print('请输入如下格式的时间%Y-%m-%d，如2017-01-01')
        return [0, 'start or end time Error']
    if endtime > time.time() * 1000:
        return [0, '结束时间超出检索范围（本日数据未存储，最大日期为今日的前一天）']
    # 判断时间参数
    find_one = b.mdb['order_lite_list'].find_one({
        'appCode': appcode,
        'status': {
            '$in': [4, 7]
        },
        'createdAt': {
            '$gte': starttime,
            '$lt': endtime
        }
    })
    if not find_one:
        print('order_name输入错误，无此字段或查无数据')
        return [0, 'order_name Error']
    # 判断type参数
    return mongodb_datas_fetch(appcode, starttime, endtime)


def mongodb_datas_fetch(appcode, starttime, endtime, datas=None):
    if datas is None:
        datas = []
    # 查询mongo数据库
    items = b.mdb['order_lite_list'].find({
        'appCode': appcode,
        'status': {
            '$in': [4, 7]
        },
        'createdAt': {
            '$gte': starttime,
            '$lt': endtime
        }
    }, {'_id': 0})
    for item in items:
        order_type = item['orderType']
        order_id = item['id']
        order_detail = b.mdb['order_{}'.format(order_type)].find_one({
            'id': order_id
        })

        order_channel = item['channel']
        order_phone = order_detail.get('phone')
        user_id = item['userId']
        order_name = item['name']
        partner_id = order_detail['partnerId']
        # partner_id = item.get('partnerId', '')
        xdtime = str(datetime.fromtimestamp(item['createdAt'] / 1000)).split('.')[0]  # yapf: disable
        order_status = item['displayStatus']
        number_status = item['status']

        data = {
            'order_type': order_type,
            'order_id': order_id,
            'order_channel': order_channel,
            'order_phone': order_phone,
            'user_id': user_id,
            'order_name': order_name,
            'xdtime': xdtime,
            'order_status': order_status,
            'number_status': number_status,
            'partner_id': partner_id,
            'appcode': appcode,
        }
        datas.append(data)
    return redis_datas_fetch(datas)


def redis_datas_fetch(datas, report_datas=None):
    if report_datas is None:
        report_datas = []
    for item in datas:
        # redis 数据库获取缺失订单价格数据字段，1为完成订单，2为退款订单
        rdb_1_datas = b.rdb_1.hgetall('p:{0}:{1}'.format(item['appcode'], item[
            'order_id']))
        rdb_2_datas = b.rdb_2.hgetall('r:{0}:{1}'.format(item['appcode'], item[
            'order_id']))

        number_status = item['number_status']
        if number_status == 4:
            pay_id = rdb_1_datas.get('id', '')
            pay_channel = rdb_1_datas.get('channel', '')
            price = ''
            if rdb_1_datas.get('price'):
                price = float(rdb_1_datas.get('price')) / 100

            zftime = ''
            if rdb_1_datas.get('createdat'):
                zftime = str(
                    datetime.fromtimestamp(
                        float(rdb_1_datas.get('createdat')) / 1000)).split(
                            '.')[0]

            payprice = ''
            if rdb_1_datas.get('payprice'):
                payprice = float(rdb_1_datas.get('payprice')) / 100

        elif number_status == 7:
            pay_id = rdb_2_datas.get('id', '')
            pay_channel = rdb_2_datas.get('channel', '')
            price = ''
            if rdb_2_datas.get('price'):
                price = float(rdb_2_datas.get('price')) / 100

            zftime = ''
            if rdb_2_datas.get('createdat'):
                zftime = str(
                    datetime.fromtimestamp(
                        float(rdb_2_datas.get('createdat')) / 1000)).split(
                            '.')[0]

            payprice = ''
            if rdb_2_datas.get('paidprice'):
                payprice = float(rdb_2_datas.get('paidprice')) / 100

        yhprice = ''
        if price and payprice:
            yhprice = price - payprice

        item['pay_id'] = pay_id
        item['pay_channel'] = pay_channel
        item['order_price'] = price
        item['zftime'] = zftime
        item['payprice'] = payprice
        item['yhprice'] = yhprice
        report_datas.append(item)
    return report_datas


def generate_report(report_datas, field_header):
    wb = Workbook()
    ws = wb.active
    ws.append(field_header)

    # 动态创建各分工作表
    items = copy.deepcopy(report_datas)
    for item in items:
        item = map_values(item)
        item = map_keys(item)
        row = ws.max_row + 1
        for col in range(1, ws.max_column + 1):
            ws.cell(
                row=row, column=col).value = item[ws.cell(
                    row=1, column=col).value]

    render_report(ws)
    # 报表完成，保存为excel文件
    wb.save("{0}_交易明细报表_{1}.xlsx".format(report_datas[0]['appcode'],
                                         len(field_header)))
    return ['file save is ok']


def render_report(ws):
    # 冻结第一行
    ws.freeze_panes = 'A2'
    for i in range(1, ws.max_column + 1):
        # 抬头第一行 设置填充色、字体格式、居中
        ws['{}1'.format(get_column_letter(i))].fill = PatternFill(
            "solid", fgColor="FEC000")
        ws['{}1'.format(get_column_letter(i))].font = Font(
            name="微软雅黑", size=12, bold=True)
        ws['{}1'.format(get_column_letter(i))].alignment = Alignment(
            horizontal='center', vertical='center')
    # 为 所有表格设置边框和字体。
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(
                row=row, column=col).border = Border(
                    left=left, right=right, top=top,
                    bottom=bottom)  # 设置单元格边框格式
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).font = Font(name="微软雅黑", size=11)


def map_values(report_data):
    map_order_type = {
        'balance': 'balance',
        'baoyang': '保养',
        'coffee': '星巴克',
        'daijia': '代驾',
        'dianying': '电影',
        'huafei': '话费',
        'huoche': '火车',
        'jiadianqingxi': '家电清洗',
        'jiadianweixiu': '家电维修',
        'jiayouka': '加油卡',
        'jipiao': '机票',
        'jiudian': '酒店',
        'liuliang': '流量',
        'menpiao': '门票',
        'paotui': '跑腿',
        'piaowu': '票务',
        'sdm': 'sdm',
        'shenghuojiaofei': '生活缴费',
        'shengxian': '生鲜',
        'waimai': '外卖',
        'xianhua': '鲜花',
        'xihu': '洗护',
        'zhuanche': '专车',
        'balance': '钱包',
    }
    map_order_channel = {
        'ele': '饿了么',
        'tieyou': '铁友',
        'dhst': '大汉三通',
        'zhenlv': '真旅',
        'ofpay': '欧飞',
        'balance': '菠萝觅',
        'linqu': '邻趣',
        'e': 'e代驾',
        'fft': '付费通',
        'sfdj': '十分到家',
        'elong': '艺龙',
        'tidy': '泰迪',
        'ai': '爱代驾',
        'ctrip': '携程',
        'lvmama': '驴妈妈',
        'tongcheng': '同程',
    }
    map_pay_channel = {
        'wx': '微信',
        'alipay': '支付宝',
        'balance': '菠萝觅',
    }
    map_appcode = {
        'boluome': '菠萝觅',
        'fft': '付费通',
        'allinpay': '通联钱包',
        'jst': '聚事通',
        'roobo': 'Roobo智能生活',
        'chubao': '触宝电话',
        'jhdp': '聚浩大屏',
        '91jincai': '91金彩',
        'mybosc': '上行快线',
        '91ala': '51返呗',
        'scity': '市民服务'
    }

    report_data['appcode'] = map_appcode[report_data['appcode']]
    report_data['order_type'] = map_order_type[report_data['order_type']]

    pay_channel = report_data['pay_channel']
    if ',' in pay_channel:
        report_data['pay_channel'] = ''
        for key in pay_channel.split(','):
            report_data['pay_channel'] += map_pay_channel[key] + ','
        report_data['pay_channel'] = report_data['pay_channel'][:-1]
    elif pay_channel:
        report_data['pay_channel'] = map_pay_channel[pay_channel]

    report_data['order_channel'] = map_order_channel[report_data[
        'order_channel'].lower()]
    return report_data


def map_keys(report_data):
    map_key = {
        'appcode': '应用名称',
        'order_id': '订单号',
        'order_type': '品类',
        'order_channel': '供应商',
        'partner_id': '供应商订单号',
        'order_phone': '手机号',
        'user_id': '用户ID',
        'order_name': '订单名称',
        'order_price': '订单金额',
        'payprice': '支付金额',
        'yhprice': '优惠金额',
        'order_status': '订单状态',
        'pay_channel': '支付方式',
        'xdtime': '下单时间',
        'zftime': '支付时间',
        'pay_id': '支付订单号',
    }
    report_data = {
        map_key[k]: v
        for k, v in report_data.items() if k in map_key
    }
    return report_data


def report_datas_fetch(appcode, start, end):
    """传入参数：

    >>>appcode='allinpay'
    >>>start='2017-01-01',
    >>>end='2017-01-31'

    返回：
    大于等于start日期零时，小于等于end日期24时的appcode名的数据报表。
    """
    appcode = str(appcode)
    start = str(start)
    end = str(end)
    # 字符化

    report_datas = parameter_verification(appcode, start, end)

    field_header_client = [
        '订单号', '品类', '供应商', '订单名称', '订单金额', '支付金额', '优惠金额', '订单状态', '支付方式',
        '下单时间', '支付时间', '支付订单号'
    ]
    field_header_admin = [
        '应用名称', '订单号', '品类', '供应商', '供应商订单号', '手机号', '用户ID', '订单名称', '订单金额',
        '支付金额', '优惠金额', '订单状态', '支付方式', '下单时间', '支付时间', '支付订单号'
    ]

    generate_report(report_datas, field_header_client)
    generate_report(report_datas, field_header_admin)
